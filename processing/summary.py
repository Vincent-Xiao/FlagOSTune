#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

DATA_LINE_RE = re.compile(
    r"^(?P<status>\S+)\s+"
    r"(?P<torch>[\d.eE+-]+)\s+"
    r"(?P<gems>[\d.eE+-]+)\s+"
    r"(?P<speedup>[\d.eE+-]+)\s+"
    r"(?P<tflops>[\d.eE+-]+)\s+"
    r"(?P<shape>\[.*\])\s*$"
)
TORCH_SIZE_RE = re.compile(r"torch\.Size\(\[(?P<dims>[^\]]+)\]\)")
SCALAR_TAIL_RE = re.compile(
    r",\s*[-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][-+]?\d+)?\s*\]$"
)
ShapeKey = tuple[object, ...]

TABLE_HEADERS = [
    "Shape (B, M, N, K)",
    "Count",
    "Torch Latency with left configuration (ms)",
    "Gems Latency with left configuration (ms)",
    "Gems Speedup with left configuration",
    "Torch Latency with right configuration (ms)",
    "Gems Latency with right configuration (ms)",
    "Gems Speedup with right configuration",
    "Speedup Gain",
]

COMPARISON_TABLE_TITLE = "Sorted by Speedup Gain"
SINGLE_CONFIG_TABLE_TITLE = "Performance Summary"
OP_SHAPE_LABELS = {
    "mul": "Inputs (kind and shapes)",
    "sparse_attention": "Shape (B, M, KV_LEN, TOPK, H, D)",
}


def parse_bool_arg(value: str) -> bool:
    normalized = value.strip().lower()
    if normalized in {"1", "true", "yes", "y", "on"}:
        return True
    if normalized in {"0", "false", "no", "n", "off"}:
        return False
    raise argparse.ArgumentTypeError(f"Invalid boolean value: {value}")


def resolve_shape_source_op(op: str) -> str:
    if op.startswith("w8a8_block_fp8_matmul"):
        return "w8a8_block_fp8_matmul"
    return op


def get_shape_label(op: str) -> str:
    return OP_SHAPE_LABELS.get(resolve_shape_source_op(op), "Shape (B, M, N, K)")


def freeze_shape_item(item):
    if isinstance(item, list):
        return tuple(freeze_shape_item(value) for value in item)
    return item


def normalize_config_shape(op: str, shape: list[object]) -> ShapeKey:
    key = tuple(freeze_shape_item(item) for item in shape)
    if resolve_shape_source_op(op) != "mul":
        return key

    kind = key[0] if key else None
    expected_length = 3 if kind == "broadcast" else 2
    if kind not in {"broadcast", "same", "scalar"} or len(key) != expected_length:
        raise ValueError(f"Invalid mul shape: {shape}")
    return key


def parse_count_map(count_yaml_path: Path, target_op: str) -> dict[ShapeKey, int]:
    target_shape_op = resolve_shape_source_op(target_op)
    if target_shape_op == "mul":
        yaml_config = yaml.safe_load(count_yaml_path.read_text(encoding="utf-8")) or {}
        count_map: dict[ShapeKey, int] = {}
        for shape in yaml_config.get("mul", {}).get("shapes", []):
            if not isinstance(shape, list) or len(shape) < 3:
                raise ValueError(f"Invalid mul count shape: {shape}")
            *shape_without_count, count = shape
            if isinstance(count, bool) or not isinstance(count, int):
                raise ValueError(f"Invalid mul count: {count}")
            key = normalize_config_shape("mul", shape_without_count)
            count_map[key] = count
        return count_map

    count_map: dict[ShapeKey, int] = {}
    current_op: str | None = None
    in_shapes = False
    current_shape: list[int] | None = None

    def flush_current_shape() -> None:
        nonlocal current_shape
        if current_op != target_shape_op or current_shape is None or len(current_shape) < 2:
            current_shape = None
            return

        *shape_key, count = current_shape
        count_map[tuple(shape_key)] = count
        current_shape = None

    with count_yaml_path.open("r", encoding="utf-8", errors="ignore") as file:
        for raw_line in file:
            line = raw_line.rstrip("\n")

            op_match = re.match(r"^([A-Za-z_][A-Za-z0-9_]*)\:$", line)
            if op_match:
                flush_current_shape()
                current_op = op_match.group(1)
                in_shapes = False
                current_shape = None
                continue

            if current_op != target_shape_op:
                continue

            if line.strip() == "shapes:":
                in_shapes = True
                continue

            if line.strip().startswith("shape_desc:"):
                flush_current_shape()
                in_shapes = False
                current_shape = None
                continue

            if not in_shapes:
                continue

            if line.startswith("  - - "):
                flush_current_shape()
                first = int(line.split("  - - ", 1)[1].strip())
                current_shape = [first]
                continue

            dim_match = re.match(r"^\s*-\s*(\d+)\s*$", line)
            if dim_match and current_shape is not None:
                current_shape.append(int(dim_match.group(1)))

    flush_current_shape()

    return count_map


def infer_bmnk_from_shape_text(shape_text: str) -> tuple[int, int, int, int] | None:
    sizes: list[tuple[int, ...]] = []
    for match in TORCH_SIZE_RE.finditer(shape_text):
        dims = tuple(
            int(token.strip())
            for token in match.group("dims").split(",")
            if token.strip()
        )
        sizes.append(dims)

    if len(sizes) < 2:
        return None

    first, second = sizes[0], sizes[1]
    if len(first) != 2 or len(second) != 2:
        return None

    m, k1 = first
    dim2_a, dim2_b = second

    if dim2_a == k1:
        return (1, m, dim2_b, k1)
    if dim2_b == k1:
        return (1, m, dim2_a, k1)

    return None


def infer_sparse_attention_shape_from_text(shape_text: str) -> tuple[int, int, int, int, int, int] | None:
    sizes: list[tuple[int, ...]] = []
    for match in TORCH_SIZE_RE.finditer(shape_text):
        dims = tuple(
            int(token.strip())
            for token in match.group("dims").split(",")
            if token.strip()
        )
        sizes.append(dims)

    if len(sizes) < 4:
        return None

    query_shape, kv_shape, _, topk_shape = sizes[:4]
    if len(query_shape) != 4 or len(kv_shape) != 3 or len(topk_shape) != 3:
        return None

    b, m, h, d = query_shape
    kv_b, kv_len, kv_d = kv_shape
    topk_b, topk_m, topk = topk_shape
    if b != kv_b or b != topk_b or m != topk_m or d != kv_d:
        return None

    return (b, m, kv_len, topk, h, d)


def infer_mul_shape_from_text(shape_text: str) -> ShapeKey | None:
    sizes = [
        tuple(
            int(token.strip())
            for token in match.group("dims").split(",")
            if token.strip()
        )
        for match in TORCH_SIZE_RE.finditer(shape_text)
    ]
    if len(sizes) == 2:
        if sizes[0] == sizes[1]:
            return "same", sizes[0]
        return "broadcast", sizes[0], sizes[1]
    if len(sizes) == 1 and SCALAR_TAIL_RE.search(shape_text):
        return "scalar", sizes[0]
    return None


def infer_shape_key_from_shape_text(
    shape_text: str, target_op: str
) -> ShapeKey | None:
    if resolve_shape_source_op(target_op) == "mul":
        return infer_mul_shape_from_text(shape_text)
    if resolve_shape_source_op(target_op) == "sparse_attention":
        return infer_sparse_attention_shape_from_text(shape_text)
    return infer_bmnk_from_shape_text(shape_text)


def shape_key_to_display(shape_key: ShapeKey, op: str) -> str:
    if resolve_shape_source_op(op) != "mul":
        return ", ".join(str(value) for value in shape_key)

    kind = shape_key[0]
    lhs_shape = shape_key[1]
    if kind == "broadcast":
        return f"broadcast: {lhs_shape} x {shape_key[2]}"
    if kind == "same":
        return f"same: {lhs_shape} x {lhs_shape}"
    return f"scalar: {lhs_shape} x scalar"


def convert_shape_to_display_and_count(
    shape_text: str,
    count_map: dict[ShapeKey, int],
    target_op: str,
    default_count: int | str = "-",
) -> tuple[str, str]:
    shape_key = infer_shape_key_from_shape_text(shape_text, target_op)
    if shape_key is None:
        return shape_text, str(default_count)

    count = count_map.get(shape_key, default_count)
    return shape_key_to_display(shape_key, target_op), str(count)


def calc_gain_percent(default_speedup: str, expand_speedup: str) -> str:
    try:
        default_val = float(default_speedup)
        expand_val = float(expand_speedup)
    except ValueError:
        return "-"

    if default_val == 0:
        return "-"

    gain = (expand_val / default_val - 1.0) * 100.0
    return f"{gain:.2f}%"


def parse_gain_value(gain_text: str) -> float:
    if not gain_text.endswith("%"):
        return float("-inf")
    try:
        return float(gain_text[:-1])
    except ValueError:
        return float("-inf")


def parse_count_value(count_text: str) -> int:
    try:
        return int(count_text)
    except ValueError:
        return -1


def parse_speedup_value(speedup_text: str) -> float:
    try:
        return float(speedup_text)
    except ValueError:
        return float("-inf")


def parse_percent_value(percent_text: str) -> float:
    if not percent_text.endswith("%"):
        raise ValueError(percent_text)
    return float(percent_text[:-1])


def get_summary_column_kinds(include_right_comparison: bool) -> list[str]:
    if include_right_comparison:
        return ["label", "count", "latency", "latency", "speedup", "latency", "latency", "speedup", "percent"]
    return ["label", "count", "latency", "latency", "speedup"]


def parse_summary_numeric(kind: str, value_text: str) -> float | None:
    try:
        if kind == "count":
            return float(int(value_text))
        if kind in {"latency", "speedup"}:
            return float(value_text)
        if kind == "percent":
            return parse_percent_value(value_text)
    except ValueError:
        return None
    return None


def format_summary_numeric(kind: str, stat_name: str, value: float) -> str:
    if kind == "count":
        if stat_name in {"Min", "Max"}:
            return str(int(round(value)))
        return f"{value:.2f}"
    if kind == "latency":
        return f"{value:.6f}"
    if kind == "speedup":
        return f"{value:.3f}"
    if kind == "percent":
        return f"{value:.2f}%"
    return "-"


def append_summary_rows(rows: list[list[str]], include_right_comparison: bool) -> list[list[str]]:
    if not rows:
        return rows

    column_kinds = get_summary_column_kinds(include_right_comparison)
    summary_specs = [
        ("Min", min),
        ("Max", max),
        ("Avg", lambda values: sum(values) / len(values)),
    ]
    summary_rows: list[list[str]] = []

    for stat_name, reducer in summary_specs:
        summary_row = [stat_name]
        for col_idx in range(1, len(column_kinds)):
            kind = column_kinds[col_idx]
            values = [
                parsed
                for row in rows
                if col_idx < len(row)
                for parsed in [parse_summary_numeric(kind, row[col_idx])]
                if parsed is not None
            ]
            if not values:
                summary_row.append("-")
                continue
            summary_value = reducer(values)
            summary_row.append(format_summary_numeric(kind, stat_name, summary_value))
        summary_rows.append(summary_row)

    return rows + summary_rows


def parse_model_yaml(model_yaml_path: Path) -> list[dict[str, object]]:
    yaml_config = yaml.safe_load(model_yaml_path.read_text(encoding="utf-8")) or {}
    if not isinstance(yaml_config, dict):
        raise ValueError(f"Invalid model YAML: {model_yaml_path}")

    blocks: list[dict[str, object]] = []
    for op, config in yaml_config.items():
        if not isinstance(config, dict):
            raise ValueError(f"Invalid config for operator '{op}'")
        shapes = config.get("shapes", [])
        if shapes is None:
            shapes = []
        if not isinstance(shapes, list):
            raise ValueError(f"Invalid shapes for operator '{op}'")
        blocks.append(
            {
                "op": op,
                "shapes": shapes,
                "shape_desc": config.get("shape_desc"),
            }
        )

    return blocks


def write_model_shapes_yaml(output_path: Path, blocks: list[dict[str, object]]) -> None:
    yaml_config: dict[str, dict[str, object]] = {}
    for block in blocks:
        op = str(block["op"])
        config: dict[str, object] = {"shapes": block["shapes"]}
        if block.get("shape_desc"):
            config["shape_desc"] = block["shape_desc"]
        yaml_config[op] = config

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as file:
        yaml.safe_dump(
            yaml_config,
            file,
            allow_unicode=True,
            sort_keys=False,
        )


def split_and_write_gain_lose_yaml(
    model_yaml_path: Path,
    gain_yaml_path: Path,
    lose_yaml_path: Path,
    op: str,
    rows_by_gain: list[list[str]],
    include_right_comparison: bool = True,
) -> tuple[int, int]:
    source_blocks = parse_model_yaml(model_yaml_path)
    display_to_gain: dict[str, float] = {}
    target_shape_op = resolve_shape_source_op(op)

    for row in rows_by_gain:
        shape_text = row[0]
        score_text = row[8] if include_right_comparison else row[4]
        if include_right_comparison:
            display_to_gain[shape_text] = parse_gain_value(score_text)
        else:
            try:
                display_to_gain[shape_text] = float(score_text)
            except ValueError:
                display_to_gain[shape_text] = float("-inf")

    gain_blocks: list[dict[str, object]] = []
    lose_blocks: list[dict[str, object]] = []
    gain_count = 0
    lose_count = 0

    for block in source_blocks:
        block_op = block["op"]
        block_shapes = block["shapes"]
        block_shape_desc = block["shape_desc"]

        if block_op != target_shape_op:
            copied_shapes = [shape.copy() for shape in block_shapes]
            gain_blocks.append(
                {
                    "op": block_op,
                    "shapes": copied_shapes,
                    "shape_desc": block_shape_desc,
                }
            )
            lose_blocks.append(
                {
                    "op": block_op,
                    "shapes": copied_shapes,
                    "shape_desc": block_shape_desc,
                }
            )
            continue

        gain_shapes: list[list[object]] = []
        lose_shapes: list[list[object]] = []
        for shape in block_shapes:
            if not shape:
                continue
            key = normalize_config_shape(str(block_op), shape)
            shape_display = shape_key_to_display(key, str(block_op))
            gain_value = display_to_gain.get(shape_display, float("-inf"))
            if include_right_comparison:
                is_gain_shape = gain_value > 0
            else:
                is_gain_shape = gain_value >= 1.0

            if is_gain_shape:
                gain_shapes.append(shape.copy())
                gain_count += 1
            else:
                lose_shapes.append(shape.copy())
                lose_count += 1

        gain_blocks.append(
            {"op": block_op, "shapes": gain_shapes, "shape_desc": block_shape_desc}
        )
        lose_blocks.append(
            {"op": block_op, "shapes": lose_shapes, "shape_desc": block_shape_desc}
        )

    write_model_shapes_yaml(gain_yaml_path, gain_blocks)
    write_model_shapes_yaml(lose_yaml_path, lose_blocks)
    return gain_count, lose_count


def append_table(
    lines: list[str],
    title: str,
    rows: list[list[str]],
    shape_label: str,
    left_report_label: str,
    right_report_label: str,
    include_right_comparison: bool,
) -> None:
    lines.append(f"## {title}")
    lines.append("")
    if include_right_comparison:
        lines.append(
            f"| {shape_label} | Count | {left_report_label} |  |  | "
            f"{right_report_label} |  |  | Speedup Gain |"
        )
        lines.append("| --- | --- | --- | --- | --- | --- | --- | --- | --- |")
        lines.append(
            f"|  |  | Torch Latency (ms) | Gems Latency (ms) | Gems Speedup | "
            f"Torch Latency (ms) | Gems Latency (ms) | Gems Speedup | {right_report_label} vs {left_report_label} |"
        )
    else:
        lines.append(f"| {shape_label} | Count | {left_report_label} |  |  |")
        lines.append("| --- | --- | --- | --- | --- |")
        lines.append("|  |  | Torch Latency (ms) | Gems Latency (ms) | Gems Speedup |")
    for row in append_summary_rows(rows, include_right_comparison):
        lines.append("| " + " | ".join(row) + " |")
    lines.append("")


def build_table_rows(
    sections: dict[str, dict[str, tuple[str, str, str]]],
    shape_order: list[str],
    count_map: dict[tuple[int, ...], int],
    target_op: str,
    default_count: int | str = "-",
    include_right_comparison: bool = True,
) -> tuple[list[list[str]], list[list[str]]]:
    table_rows: list[list[str]] = []

    for shape in shape_order:
        left_metrics = sections["left"].get(shape, ("-", "-", "-"))
        shape_bmnk, count = convert_shape_to_display_and_count(
            shape,
            count_map,
            target_op,
            default_count,
        )
        row = [shape_bmnk, count, left_metrics[0], left_metrics[1], left_metrics[2]]

        if include_right_comparison:
            right_metrics = sections["right"].get(shape, ("-", "-", "-"))
            gain = calc_gain_percent(left_metrics[2], right_metrics[2])
            row.extend([right_metrics[0], right_metrics[1], right_metrics[2], gain])

        table_rows.append(row)

    if include_right_comparison:
        rows_by_gain = sorted(table_rows, key=lambda row: parse_gain_value(row[8]), reverse=True)
    else:
        rows_by_gain = sorted(table_rows, key=lambda row: parse_speedup_value(row[4]), reverse=True)
    rows_by_count = sorted(table_rows, key=lambda row: parse_count_value(row[1]), reverse=True)
    return rows_by_gain, rows_by_count


def write_excel_report(
    xlsx_path: Path,
    rows_by_gain: list[list[str]],
    rows_by_count: list[list[str]],
    shape_label: str,
    left_report_label: str,
    right_report_label: str,
    include_right_comparison: bool = True,
    include_count_sheet: bool = True,
) -> None:
    def style_sheet(ws) -> None:
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = center_alignment

        for row_idx in (1, 2):
            for cell in ws[row_idx]:
                cell.font = header_font

        for col_idx in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[column_letter]:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_len:
                    max_len = len(value)

            adjusted_width = max(12, min(max_len + 2, 60))
            ws.column_dimensions[column_letter].width = adjusted_width

    def write_sheet(ws, rows: list[list[str]]) -> None:
        if include_right_comparison:
            ws.append([
                shape_label,
                "Count",
                left_report_label,
                "",
                "",
                right_report_label,
                "",
                "",
                "Speedup Gain",
            ])
            ws.append([
                "",
                "",
                "Torch Latency (ms)",
                "Gems Latency (ms)",
                "Gems Speedup",
                "Torch Latency (ms)",
                "Gems Latency (ms)",
                "Gems Speedup",
                f"{right_report_label} vs {left_report_label}",
            ])

            ws.merge_cells("A1:A2")
            ws.merge_cells("B1:B2")
            ws.merge_cells("C1:E1")
            ws.merge_cells("F1:H1")
            ws.merge_cells("I1:I2")
        else:
            ws.append([
                shape_label,
                "Count",
                left_report_label,
                "",
                "",
            ])
            ws.append([
                "",
                "",
                "Torch Latency (ms)",
                "Gems Latency (ms)",
                "Gems Speedup",
            ])

            ws.merge_cells("A1:A2")
            ws.merge_cells("B1:B2")
            ws.merge_cells("C1:E1")

        for row in append_summary_rows(rows, include_right_comparison):
            ws.append(row)

        style_sheet(ws)

    workbook = Workbook()

    ws_gain = workbook.active
    ws_gain.title = COMPARISON_TABLE_TITLE if include_right_comparison else SINGLE_CONFIG_TABLE_TITLE
    write_sheet(ws_gain, rows_by_gain)

    if include_count_sheet:
        ws_count = workbook.create_sheet(title="Sorted by Count")
        write_sheet(ws_count, rows_by_count)

    workbook.save(xlsx_path)


def parse_log(
    log_path: Path,
    left_stage_label: str = "default",
    right_stage_label: str = "expand",
) -> tuple[dict[str, dict[str, tuple[str, str, str]]], list[str]]:
    left_info_re = re.compile(rf"\[INFO\].*with {re.escape(left_stage_label)} configuration\.")
    right_info_re = re.compile(rf"\[INFO\].*with {re.escape(right_stage_label)} configuration\.")
    sections: dict[str, dict[str, tuple[str, str, str]]] = {"left": {}, "right": {}}
    shape_order: list[str] = []
    current_section: str | None = None

    with log_path.open("r", encoding="utf-8", errors="ignore") as file:
        for raw_line in file:
            line = raw_line.strip()
            if not line:
                continue

            if left_info_re.search(line):
                current_section = "left"
                continue
            if right_info_re.search(line):
                current_section = "right"
                continue

            if current_section is None:
                continue

            match = DATA_LINE_RE.match(line)
            if not match:
                continue

            shape = match.group("shape")
            metrics = (match.group("torch"), match.group("gems"), match.group("speedup"))
            sections[current_section][shape] = metrics

            if shape not in shape_order:
                shape_order.append(shape)

    return sections, shape_order


def build_report_content(
    model: str,
    op: str,
    rows_by_gain: list[list[str]],
    rows_by_count: list[list[str]],
    count_yaml_exists: bool,
    left_report_label: str,
    right_report_label: str,
    include_right_comparison: bool,
) -> str:
    lines: list[str] = []
    shape_label = get_shape_label(op)
    lines.append(f"# Performance Summary: {model} / {op}")
    lines.append("")
    lines.append(f"- Source log: `log/flagtune/{model}/{op}/pretune/pretune.log`")
    if include_right_comparison:
        lines.append(f"- Compare: `{left_report_label}` vs `{right_report_label}`")
    else:
        lines.append(f"- Configuration: `{left_report_label}`")
    if count_yaml_exists:
        lines.append(f"- Count reference: `FlagTune/shape-config/{model}_count.yaml`")
    else:
        lines.append("- Count reference: missing, all counts fallback to `1`")
    lines.append(f"- Rows: {len(rows_by_gain)}")
    lines.append("")

    primary_title = COMPARISON_TABLE_TITLE if include_right_comparison else SINGLE_CONFIG_TABLE_TITLE
    append_table(
        lines,
        primary_title,
        rows_by_gain,
        shape_label,
        left_report_label,
        right_report_label,
        include_right_comparison,
    )
    if count_yaml_exists:
        append_table(
            lines,
            "Sorted by Count",
            rows_by_count,
            shape_label,
            left_report_label,
            right_report_label,
            include_right_comparison,
        )

    lines.append("")
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate markdown summary from pretune log",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--model", default="qwen3.5", help="Model name")
    parser.add_argument("--op", default="mm", help="Operator name")
    parser.add_argument("--output-suffix", default="", help="Suffix appended to report filenames, for example _master")
    parser.add_argument("--left-stage-label", default="default", help="Log stage name for the left-side comparison")
    parser.add_argument("--right-stage-label", default="expand", help="Log stage name for the right-side comparison")
    parser.add_argument("--left-report-label", default="Default Configuration", help="Report column title for the left-side comparison")
    parser.add_argument("--right-report-label", default="Expand Configuration", help="Report column title for the right-side comparison")
    parser.add_argument(
        "--include-right-comparison",
        type=parse_bool_arg,
        default=True,
        help="Whether to include the right-side comparison columns in the report",
    )
    args = parser.parse_args()

    repo_root = Path(__file__).resolve().parent.parent.parent
    flagtune_dir = Path(__file__).resolve().parent.parent
    log_path = repo_root / "log" / "flagtune" / args.model / args.op / "pretune" / "pretune.log"
    count_yaml_path = flagtune_dir / "shape-config" / f"{args.model}_count.yaml"
    model_yaml_path = flagtune_dir / "shape-config" / f"{args.model}.yaml"
    gain_yaml_path = flagtune_dir / "shape-config" / f"{args.model}_gain.yaml"
    lose_yaml_path = flagtune_dir / "shape-config" / f"{args.model}_lose.yaml"
    output_path = flagtune_dir / "reports" / f"{args.model}_{args.op}{args.output_suffix}.md"
    output_xlsx_path = flagtune_dir / "reports" / f"{args.model}_{args.op}{args.output_suffix}.xlsx"

    if not log_path.exists():
        raise FileNotFoundError(f"Log file not found: {log_path}")
    if not model_yaml_path.exists():
        raise FileNotFoundError(f"Model yaml not found: {model_yaml_path}")

    sections, shape_order = parse_log(log_path, args.left_stage_label, args.right_stage_label)
    count_yaml_exists = count_yaml_path.exists()
    count_map = parse_count_map(count_yaml_path, args.op) if count_yaml_exists else {}
    shape_label = get_shape_label(args.op)
    if not shape_order:
        raise ValueError("No performance rows were parsed from the log file")

    default_count = 1 if not count_yaml_exists else "-"
    rows_by_gain, rows_by_count = build_table_rows(
        sections,
        shape_order,
        count_map,
        args.op,
        default_count,
        include_right_comparison=args.include_right_comparison,
    )

    report = build_report_content(
        args.model,
        args.op,
        rows_by_gain,
        rows_by_count,
        count_yaml_exists,
        args.left_report_label,
        args.right_report_label,
        args.include_right_comparison,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(report, encoding="utf-8")
    write_excel_report(
        output_xlsx_path,
        rows_by_gain,
        rows_by_count,
        shape_label,
        args.left_report_label,
        args.right_report_label,
        include_right_comparison=args.include_right_comparison,
        include_count_sheet=count_yaml_exists,
    )

    print(f"Generated report: {output_path}")
    print(f"Generated report: {output_xlsx_path}")
    gain_count, lose_count = split_and_write_gain_lose_yaml(
        model_yaml_path,
        gain_yaml_path,
        lose_yaml_path,
        args.op,
        rows_by_gain,
        include_right_comparison=args.include_right_comparison,
    )
    print(f"Generated gain yaml: {gain_yaml_path} (shapes={gain_count})")
    print(f"Generated lose yaml: {lose_yaml_path} (shapes={lose_count})")
    print(f"Rows: {len(rows_by_gain)}")


if __name__ == "__main__":
    main()
