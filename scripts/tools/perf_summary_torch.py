#!/usr/bin/env python3
"""汇总各模型 perf_analysis_torch.md 的 TOP CUDA/FlagGems Kernel 占比表。"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

try:
    from openpyxl import Workbook
except ImportError:  # pragma: no cover
    Workbook = None


CUDA_SECTION_TITLES: Sequence[str] = (
    "## CUDA kernel (sorted by total time)",
    "## CUDA kernel（按总时间排序）",
)
FLAGGEMS_SECTION_TITLES: Sequence[str] = (
    "## FlagGems kernel (sorted by total time)",
    "## FlagGems kernel（按总时间排序）",
)


def parse_percent(percent_text: str) -> float:
    text = percent_text.strip()
    if text.endswith("%"):
        text = text[:-1]
    return float(text) / 100.0


def split_markdown_row(line: str) -> List[str]:
    text = line.strip()
    if text.startswith("|"):
        text = text[1:]
    if text.endswith("|"):
        text = text[:-1]

    parts: List[str] = []
    cell_chars: List[str] = []
    escaped = False
    for ch in text:
        if escaped:
            cell_chars.append(ch)
            escaped = False
            continue
        if ch == "\\":
            escaped = True
            continue
        if ch == "|":
            parts.append("".join(cell_chars).strip())
            cell_chars = []
            continue
        cell_chars.append(ch)

    if escaped:
        cell_chars.append("\\")
    parts.append("".join(cell_chars).strip())
    return parts


def find_table_rows(md_text: str, section_titles: Sequence[str]) -> List[List[str]]:
    lines = md_text.splitlines()

    section_start = -1
    for idx, line in enumerate(lines):
        if line.strip() in section_titles:
            section_start = idx
            break

    if section_start < 0:
        return []

    header_idx = -1
    for idx in range(section_start + 1, len(lines)):
        stripped = lines[idx].strip()
        if stripped.startswith("| 框架算子名 |"):
            header_idx = idx
            break
        if stripped.startswith("## "):
            return []

    if header_idx < 0:
        return []

    table_rows: List[List[str]] = []
    for idx in range(header_idx + 2, len(lines)):
        line = lines[idx].strip()
        if not line or line.startswith("## ") or not line.startswith("|"):
            break
        cells = split_markdown_row(line)
        if len(cells) < 6:
            continue
        table_rows.append(cells)

    return table_rows


def parse_model_section_pct(report_path: Path, section_titles: Sequence[str], threshold: float) -> Dict[str, float]:
    content = report_path.read_text(encoding="utf-8")
    rows = find_table_rows(content, section_titles)

    op_to_pct: Dict[str, float] = {}
    for row in rows:
        framework_op = row[0]
        pct_text = row[5]
        try:
            pct = parse_percent(pct_text)
        except ValueError:
            continue
        if pct >= threshold:
            op_to_pct[framework_op] = max(op_to_pct.get(framework_op, 0.0), pct)

    return op_to_pct


def collect_model_reports(reports_root: Path) -> List[Tuple[str, Path]]:
    model_reports: List[Tuple[str, Path]] = []
    for model_dir in sorted(reports_root.iterdir(), key=lambda p: p.name.lower()):
        if not model_dir.is_dir():
            continue
        report_path = model_dir / "perf_analysis_torch.md"
        if report_path.exists():
            model_reports.append((model_dir.name, report_path))
    return model_reports


def fmt_pct(value: float) -> str:
    return f"{value * 100:.2f}%"


def build_table_header_rows(model_to_data: Dict[str, Dict[str, float]]) -> Tuple[List[str], List[List[str]]]:
    models = sorted(model_to_data.keys(), key=str.lower)

    op_set = set()
    for data in model_to_data.values():
        op_set.update(data.keys())

    op_avg_score: Dict[str, float] = {}
    for op in op_set:
        total_value = 0.0
        for model in models:
            value = model_to_data[model].get(op)
            if value is not None:
                total_value += value
        op_avg_score[op] = (total_value / len(models)) if models else 0.0

    ordered_ops = sorted(op_set, key=lambda op: (op_avg_score[op], op), reverse=True)

    header = ["序号", "框架算子名"] + models + ["占比平均值"]
    rows: List[List[str]] = []
    for idx, op in enumerate(ordered_ops, start=1):
        row = [str(idx), op]
        total_value = 0.0
        for model in models:
            value = model_to_data[model].get(op)
            if value is not None:
                total_value += value
            row.append(fmt_pct(value) if value is not None else "-")

        avg_value = total_value / len(models) if models else 0.0
        row.append(fmt_pct(avg_value))
        rows.append(row)

    return header, rows


def build_top_table_markdown(title: str, model_to_data: Dict[str, Dict[str, float]]) -> List[str]:
    header, rows = build_table_header_rows(model_to_data)

    out: List[str] = [f"## {title}", ""]
    out.append("| " + " | ".join(header) + " |")
    out.append("|" + "|".join(["---"] * len(header)) + "|")
    for row in rows:
        out.append("| " + " | ".join(row) + " |")
    out.append("")
    return out


def build_summary_markdown(
    model_to_cuda_data: Dict[str, Dict[str, float]],
    model_to_flaggems_data: Dict[str, Dict[str, float]],
) -> str:
    out: List[str] = ["# Torch Profiler Kernel 汇总报告", ""]
    out.extend(build_top_table_markdown("TOP CUDA Kernel 汇总", model_to_cuda_data))
    out.extend(build_top_table_markdown("TOP FlagGems Kernel 汇总", model_to_flaggems_data))
    return "\n".join(out)


def write_summary_excel(
    excel_path: Path,
    model_to_cuda_data: Dict[str, Dict[str, float]],
    model_to_flaggems_data: Dict[str, Dict[str, float]],
) -> None:
    if Workbook is None:
        raise SystemExit("缺少依赖 openpyxl，请先安装：pip install openpyxl")

    wb = Workbook()
    ws_cuda = wb.active
    ws_cuda.title = "TOP CUDA Kernel"
    ws_flaggems = wb.create_sheet(title="TOP FlagGems Kernel")

    def write_table(ws, title: str, model_to_data: Dict[str, Dict[str, float]]) -> None:
        header, rows = build_table_header_rows(model_to_data)

        row_idx = 1
        ws.cell(row=row_idx, column=1, value=title)
        row_idx += 1

        for col_idx, col_name in enumerate(header, start=1):
            ws.cell(row=row_idx, column=col_idx, value=col_name)
        row_idx += 1

        for row in rows:
            for col_idx, cell in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=cell)
            row_idx += 1

    write_table(ws_cuda, "TOP CUDA Kernel 汇总", model_to_cuda_data)
    write_table(ws_flaggems, "TOP FlagGems Kernel 汇总", model_to_flaggems_data)

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="汇总各模型 perf_analysis_torch.md 的 TOP CUDA/FlagGems Kernel 占比")
    parser.add_argument(
        "--reports_root",
        type=str,
        default="reports",
        help="报告根目录，默认 reports",
    )
    parser.add_argument(
        "--output",
        type=str,
        default="reports/perf_summary_torch.md",
        help="汇总输出文件路径，默认 reports/perf_summary_torch.md",
    )
    parser.add_argument(
        "--excel_output",
        type=str,
        default="reports/perf_summary_torch.xlsx",
        help="Excel 汇总输出文件路径，默认 reports/perf_summary_torch.xlsx",
    )
    parser.add_argument(
        "--threshold",
        type=float,
        default=0.0,
        help="占比阈值，默认 0.0（不过滤；0.01 表示 1%%）",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    reports_root = Path(args.reports_root)
    output_path = Path(args.output)
    excel_output_path = Path(args.excel_output)
    threshold = args.threshold

    if threshold < 0:
        raise SystemExit("--threshold 必须 >= 0")

    if not reports_root.exists() or not reports_root.is_dir():
        raise SystemExit(f"报告目录不存在: {reports_root}")

    model_reports = collect_model_reports(reports_root)
    if not model_reports:
        raise SystemExit(f"在 {reports_root} 下未找到任何 perf_analysis_torch.md")

    model_to_cuda_data: Dict[str, Dict[str, float]] = {}
    model_to_flaggems_data: Dict[str, Dict[str, float]] = {}
    for model_name, report_path in model_reports:
        model_to_cuda_data[model_name] = parse_model_section_pct(
            report_path,
            section_titles=CUDA_SECTION_TITLES,
            threshold=threshold,
        )
        model_to_flaggems_data[model_name] = parse_model_section_pct(
            report_path,
            section_titles=FLAGGEMS_SECTION_TITLES,
            threshold=threshold,
        )

    summary = build_summary_markdown(model_to_cuda_data, model_to_flaggems_data)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(summary, encoding="utf-8")
    write_summary_excel(excel_output_path, model_to_cuda_data, model_to_flaggems_data)

    print(f"已生成汇总: {output_path}")
    print(f"已生成 Excel: {excel_output_path}")
    print(f"模型数量: {len(model_reports)}")


if __name__ == "__main__":
    main()
